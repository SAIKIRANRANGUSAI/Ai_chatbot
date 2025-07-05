from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count
from django.db import IntegrityError
from django.core.paginator import Paginator
from datetime import timedelta
import json
import logging
from chatbot.models import Lead, SupportTicket
from .models import CustomAdmin
from django.db.models import Q
logger = logging.getLogger(__name__)
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Q
from chatbot.models import Lead, SupportTicket

# --------------------------
# Session Auth Helpers
# --------------------------

def get_logged_in_admin(request):
    admin_id = request.session.get('admin_user_id')
    if not admin_id:
        return None
    try:
        return CustomAdmin.objects.get(id=admin_id)
    except CustomAdmin.DoesNotExist:
        return None

# --------------------------
# Authentication Views
# --------------------------

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        try:
            user = CustomAdmin.objects.get(username=username)
            if not user.check_password(password):
                raise CustomAdmin.DoesNotExist
            request.session['admin_user_id'] = user.id
            return redirect('/core/dashboard/')
        except CustomAdmin.DoesNotExist:
            logger.warning(f"Failed login attempt for username: {username}")
            return render(request, "core/login.html", {
                "error": "Invalid username or password"
            })
    return render(request, "core/login.html")

def logout_view(request):
    request.session.flush()
    return redirect('/core/login/')

# --------------------------
# Dashboard View
# --------------------------

def dashboard(request):
    admin_user = get_logged_in_admin(request)
    if not admin_user:
        return redirect('/core/login/')
    if not admin_user.has_dashboard_access:
        logger.warning(f"Access denied for user {admin_user.username} to dashboard")
        return HttpResponse("Access denied", status=403)

    now = timezone.localtime(timezone.now())
    start_of_today = now.replace(hour=0, minute=0, second=0, microsecond=0)
    end_of_today = now.replace(hour=23, minute=59, second=59, microsecond=999999)

    # Get range from URL query (?range=7, 30, 365)
    try:
        range_days = int(request.GET.get('range', 7))
    except ValueError:
        range_days = 7

    def get_last_days_counts(model, days):
        results = []
        for i in range(days - 1, -1, -1):
            date = now - timedelta(days=i)
            start = date.replace(hour=0, minute=0, second=0, microsecond=0)
            end = date.replace(hour=23, minute=59, second=59, microsecond=999999)
            count = model.objects.filter(created_at__range=(start, end)).count()
            results.append({
                "date": date.strftime("%Y-%m-%d"),
                "count": count
            })
        return results

    lead_data = get_last_days_counts(Lead, range_days)
    support_data = get_last_days_counts(SupportTicket, range_days)

    leads_today = Lead.objects.filter(created_at__range=(start_of_today, end_of_today)).count()
    support_today = SupportTicket.objects.filter(created_at__range=(start_of_today, end_of_today)).count()

    # Placeholder: Replace below with actual Chat/Conversation model if exists
    chats_today_count = leads_today + support_today
    total_conversation_count = Lead.objects.count() + SupportTicket.objects.count()

    return render(request, 'core/dashboard.html', {
        'leads_today': leads_today,
        'support_today': support_today,
        'total_leads': Lead.objects.count(),
        'total_support': SupportTicket.objects.count(),
        'lead_data': json.dumps(lead_data),
        'support_data': json.dumps(support_data),
        'admin_user': admin_user,
        'recent_leads': Lead.objects.order_by('-created_at')[:5],
        'recent_supports': SupportTicket.objects.order_by('-created_at')[:5],
        'chats_today': chats_today_count,
        'total_conversations': total_conversation_count,
        'selected_range': range_days,
    })

# --------------------------
# Leads and Support Views
# --------------------------



def leads_table(request):
    admin_user = get_logged_in_admin(request)
    if not admin_user:
        return redirect('/core/login/')
    if not admin_user.can_view_leads:
        return HttpResponse("Access denied", status=403)

    query = request.GET.get('q', '').strip()
    leads_list = Lead.objects.all()

    if query:
        leads_list = leads_list.filter(
            Q(name__icontains=query) |
            Q(phone__icontains=query) |
            Q(email__icontains=query) |
            Q(service__icontains=query) |
            Q(requirement__icontains=query) |
            Q(budget__icontains=query)
        )

    if request.GET.get('export') == 'excel':
        return export_leads_to_excel(leads_list)

    leads_list = leads_list.order_by('-created_at')
    paginator = Paginator(leads_list, 10)
    page_number = request.GET.get('page')
    leads = paginator.get_page(page_number)

    return render(request, "core/leads_table.html", {
        "leads": leads,
        "admin_user": admin_user
    })


def support_table(request):
    admin_user = get_logged_in_admin(request)
    if not admin_user:
        return redirect('/core/login/')
    if not admin_user.can_view_support:
        return HttpResponse("Access denied", status=403)

    query = request.GET.get('q', '').strip()
    tickets_list = SupportTicket.objects.all()

    if query:
        tickets_list = tickets_list.filter(
            Q(name__icontains=query) |
            Q(phone__icontains=query) |
            Q(domain__icontains=query) |
            Q(issue__icontains=query) |
            Q(browser__icontains=query)
        )

    if request.GET.get('export') == 'excel':
        return export_support_to_excel(tickets_list)

    tickets_list = tickets_list.order_by('-created_at')
    paginator = Paginator(tickets_list, 10)
    page_number = request.GET.get('page')
    tickets = paginator.get_page(page_number)

    return render(request, "core/support_table.html", {
        "tickets": tickets,
        "admin_user": admin_user
    })



# --------------------------
# Admin Role Management Views
# --------------------------

def admin_roles_view(request):
    admin_user = get_logged_in_admin(request)
    if not admin_user:
        return redirect('/core/login/')
    if not admin_user.can_manage_roles:
        return HttpResponse("Access denied", status=403)

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        can_view_leads = 'can_view_leads' in request.POST
        can_view_support = 'can_view_support' in request.POST
        has_dashboard_access = 'has_dashboard_access' in request.POST
        can_manage_roles = 'can_manage_roles' in request.POST

        if not username or not password:
            messages.error(request, "Username and password are required.")
        else:
            try:
                new_admin = CustomAdmin(
                    username=username,
                    can_view_leads=can_view_leads,
                    can_view_support=can_view_support,
                    has_dashboard_access=has_dashboard_access,
                    can_manage_roles=can_manage_roles,
                )
                new_admin.set_password(password)
                new_admin.save()
                messages.success(request, f"Sub-admin '{username}' created successfully.")
            except IntegrityError:
                messages.error(request, "Username already exists.")

    admins = CustomAdmin.objects.all().order_by('-id')
    return render(request, "core/admin_roles.html", {
        "admins": admins,
        "admin_user": admin_user
    })

def update_admin_role(request, admin_id):
    admin_user = get_logged_in_admin(request)
    if not admin_user:
        return redirect('/core/login/')
    if not admin_user.can_manage_roles:
        return HttpResponse("Access denied", status=403)

    try:
        target = CustomAdmin.objects.get(id=admin_id)
    except CustomAdmin.DoesNotExist:
        return HttpResponse("Admin not found", status=404)

    if request.method == 'POST':
        target.has_dashboard_access = 'has_dashboard_access' in request.POST
        target.can_view_leads = 'can_view_leads' in request.POST
        target.can_view_support = 'can_view_support' in request.POST
        target.can_manage_roles = 'can_manage_roles' in request.POST
        target.save()
        messages.success(request, f"Admin '{target.username}' updated successfully.")
        return redirect('/core/admin/roles/')

    return render(request, "core/edit_admin_role.html", {
        "admin": target,
        "admin_user": admin_user
    })

# --------------------------
# Change Password View
# --------------------------

def change_password_view(request):
    admin_user = get_logged_in_admin(request)
    if not admin_user:
        return redirect('/core/login/')

    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        if not admin_user.check_password(old_password):
            messages.error(request, "Old password is incorrect.")
        elif new_password != confirm_password:
            messages.error(request, "New passwords do not match.")
        else:
            admin_user.set_password(new_password)
            admin_user.save()
            messages.success(request, "Password changed successfully.")
            return redirect('/core/change-password/')

    return render(request, "core/change_password.html", {
        "admin_user": admin_user
    })

def export_leads_to_excel(leads_queryset):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Leads"

    headers = ['Name', 'Phone', 'Email', 'Service', 'Type', 'Pages', 'CMS', 'Payment', 'Budget', 'Requirement', 'Created At']
    sheet.append(headers)

    for lead in leads_queryset:
        sheet.append([
            lead.name,
            lead.phone,
            lead.email,
            lead.service,
            lead.type,
            lead.pages,
            lead.cms,
            lead.payment,
            lead.budget,
            lead.requirement,
            lead.created_at.strftime('%Y-%m-%d %H:%M'),
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=leads.xlsx'
    workbook.save(response)
    return response


def export_support_to_excel(tickets_queryset):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Support Tickets"

    headers = ['Name', 'Phone', 'Domain', 'Issue', 'Browser', 'IP', 'Created At']
    sheet.append(headers)

    for ticket in tickets_queryset:
        sheet.append([
            ticket.name,
            ticket.phone,
            ticket.domain,
            ticket.issue,
            ticket.browser,
            ticket.ip,
            ticket.created_at.strftime('%Y-%m-%d %H:%M'),
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=support_tickets.xlsx'
    workbook.save(response)
    return response
